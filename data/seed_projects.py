#!/usr/bin/env python3
"""seed_projects.py

Read .xlsm stock files and populate real_estate_companies, projects,
and units in Supabase via the PostgREST REST API.

Usage:
    cd data/
    python seed_projects.py
"""

import io
import json
import os
import re
import sys
import unicodedata
from collections import Counter

# Force UTF-8 output on Windows
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl
import requests

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ENV_FILE   = os.path.join(SCRIPT_DIR, "..", "nextjs", ".env.local")

# ── Helpers ───────────────────────────────────────────────────────────────────

def load_env(path: str) -> dict:
    env = {}
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                k, _, v = line.partition("=")
                env[k.strip()] = v.strip()
    return env


def normalize(s) -> str:
    """Lowercase + strip + remove combining accents (for header matching)."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    return "".join(c for c in s if not unicodedata.combining(c))


def to_decimal(v):
    if v is None:
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def to_int(v):
    f = to_decimal(v)
    if f is None:
        return None
    return int(round(f))


# ── Status mapping ────────────────────────────────────────────────────────────
STATUS_MAP = {
    "disponible":  "available",
    "disponibles": "available",
    "bloqueado":   "blocked",
    "bloqueada":   "blocked",
    "sin abono":   "sin_abono",
    "sin_abono":   "sin_abono",
}

# ── Column header → DB field ──────────────────────────────────────────────────
HEADER_TO_FIELD = {
    "estado":                          "status",
    "n° depto.":                       "unit_number",
    "n depto.":                        "unit_number",
    "orientacion":                     "orientation",
    "tipologia":                       "typology",
    "tipo":                            "unit_type",
    "sup. util":                       "surface_useful",
    "sup. terraza":                    "surface_terrace",
    "sup. pond.":                      "surface_weighted",
    "sup. total":                      "surface_total",
    "p.lista":                         "list_price",
    "descuento":                       "discount",
    "p.final":                         "final_price",
    "p. de escrituracion":             "deed_price",
    "uf/m2":                           "uf_per_m2",
    "est.":                            "parking",
    "estacionamiento":                 "parking",
    "bod.":                            "storage",
    "bodega":                          "storage",
    "bono":                            "bonus_percentage",
    "bono %":                          "bonus_percentage",
    "bono uf":                         "bonus_uf",
    "credito hip. maximo":             "mortgage_max_percentage",
    "credito hip. max":                "mortgage_max_percentage",
    "credito hipotecario max":         "mortgage_max_percentage",
    "credito hipotecario maxim":       "mortgage_max_percentage",
    "credito hip. uf":                 "mortgage_max_uf",
    "credito hipotecario uf":          "mortgage_max_uf",
    "pie (%)":                         "pie_percentage",
    "pie (uf)":                        "pie_uf",
    "pie uf":                          "pie_uf",
    "pie ($)":                         "pie_clp",
    "n° cuotas cheques":               "installments_plan1",
    "n cuotas cheques":                "installments_plan1",
    "cuota cheque":                    "installments_plan1",
    "cuotas cheque + cuoton":          "installments_plan2",
    "cuotas tc":                       "installments_plan2",
    "n° cuotas tc":                    "installments_plan2",
    "n cuotas tc":                     "installments_plan2",
    # "cuota mensual $" → positional (first=plan1, second=plan2)
    "cuota mensual tc":                "monthly_payment_plan2",
    "arriendo":                        "rent_estimate",
    "dividendo":                       "mortgage_payment",
    "flujo":                           "cash_flow",
    "renta minima":                    "min_income",
}

INT_FIELDS     = {"parking", "storage", "installments_plan1", "installments_plan2"}
SKIP_NORMALIZED = {"planta"}

# ── Projects configuration ────────────────────────────────────────────────────
COMPANIES = [
    {
        "name":         "Delabase",
        "display_name": "Delabase",
        "file":         "Stock Delabase.xlsm",
        "projects": [
            {"sheet": "Volcan - Villarrica",             "name": "Volcán Villarrica",        "slug": "volcan-villarrica",        "commune": "Villarrica"},
            {"sheet": "Biaut - La Cisterna",             "name": "Biaut La Cisterna",        "slug": "biaut-la-cisterna",        "commune": "La Cisterna"},
            {"sheet": "Carvajal - La Cisterna",          "name": "Carvajal La Cisterna",     "slug": "carvajal-la-cisterna",     "commune": "La Cisterna"},
            {"sheet": "Carvajal - La Cisterna (Ext)",    "name": "Carvajal La Cisterna Ext", "slug": "carvajal-la-cisterna-ext", "commune": "La Cisterna"},
            {"sheet": "Don Claudio - La Cisterna",       "name": "Don Claudio La Cisterna",  "slug": "don-claudio-la-cisterna",  "commune": "La Cisterna"},
            {"sheet": "Don Claudio - La Cisterna (Ext)", "name": "Don Claudio La Cisterna Ext", "slug": "don-claudio-la-cisterna-ext", "commune": "La Cisterna"},
        ],
    },
    {
        "name":         "Metra",
        "display_name": "Metra",
        "file":         "Stock Metra.xlsm",
        "projects": [
            {"sheet": "Macul View - Macul",  "name": "Macul View", "slug": "macul-view", "commune": "Macul"},
            {"sheet": "Ñuñoa View - Ñuñoa", "name": "Ñuñoa View", "slug": "nunoa-view", "commune": "Ñuñoa"},
        ],
    },
    {
        "name":         "Paz Futura",
        "display_name": "Paz Futura",
        "file":         "Stock Paz - Futura.xlsm",
        "projects": [
            {"sheet": "Atelier - Ñuñoa",            "name": "Atelier",         "slug": "atelier",         "commune": "Ñuñoa"},
            {"sheet": "Mercado Serrano - Stgo",      "name": "Mercado Serrano", "slug": "mercado-serrano", "commune": "Santiago"},
            {"sheet": "Plaza Lira - Stgo Centro",    "name": "Plaza Lira",      "slug": "plaza-lira",      "commune": "Santiago"},
            {"sheet": "Seminario 2 - Ñuñoa",         "name": "Seminario 2",     "slug": "seminario-2",     "commune": "Ñuñoa"},
            {"sheet": "San Francisco - Stgo Centro", "name": "San Francisco",   "slug": "san-francisco",   "commune": "Santiago"},
        ],
    },
]

# ── Excel parsing ─────────────────────────────────────────────────────────────

def build_col_map(header_row: tuple) -> dict:
    """
    Returns {db_field: col_index}.
    "cuota mensual $" appears twice → first=monthly_payment_plan1, second=plan2.
    """
    col_map: dict = {}
    cuota_mensual_count = 0

    for idx, cell in enumerate(header_row):
        n = normalize(cell)
        if not n or n in SKIP_NORMALIZED:
            continue

        if n == "cuota mensual $":
            cuota_mensual_count += 1
            field = "monthly_payment_plan1" if cuota_mensual_count == 1 else "monthly_payment_plan2"
            col_map.setdefault(field, idx)
            continue

        field = HEADER_TO_FIELD.get(n)
        if field and field not in col_map:
            col_map[field] = idx

    return col_map


def map_row(row: tuple, col_map: dict, headers: list):
    """Convert one data row to a DB-ready dict. Returns None for empty rows."""
    if all(v is None for v in row):
        return None

    unit_num_idx = col_map.get("unit_number")
    if unit_num_idx is None or unit_num_idx >= len(row) or row[unit_num_idx] is None:
        return None

    status_idx = col_map.get("status")
    if status_idx is not None and (status_idx >= len(row) or row[status_idx] is None):
        return None

    # Build raw_data (handle duplicate header names by appending index)
    raw: dict = {}
    seen: dict = {}
    for idx, val in enumerate(row):
        h = headers[idx] if idx < len(headers) else None
        key = h if h else f"col_{idx}"
        if key in seen:
            seen[key] += 1
            key = f"{key}_{seen[key]}"
        else:
            seen[key] = 0
        raw[key] = round(val, 8) if isinstance(val, float) else val

    unit: dict = {"raw_data": raw}

    for field, idx in col_map.items():
        if idx >= len(row):
            continue
        val = row[idx]

        if field == "status":
            raw_status = str(val).strip() if val is not None else ""
            unit["original_status"] = raw_status
            unit["status"] = STATUS_MAP.get(normalize(raw_status), "blocked")
        elif field == "unit_number":
            unit["unit_number"] = str(val).strip() if val is not None else None
        elif field in INT_FIELDS:
            unit[field] = to_int(val)
        elif field in {"orientation", "typology", "unit_type"}:
            unit[field] = str(val).strip() if val is not None else None
        else:
            unit[field] = to_decimal(val)

    if "status" not in unit:
        unit["status"] = "blocked"

    return unit


def find_sheet(wb, sheet_name: str):
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    norm_target = normalize(sheet_name)
    for sn in wb.sheetnames:
        if normalize(sn) == norm_target:
            print(f"    [sheet fallback] '{sheet_name}' → '{sn}'")
            return wb[sn]
    return None


def read_sheet(wb, sheet_name: str, source_file: str) -> list:
    ws = find_sheet(wb, sheet_name)
    if ws is None:
        print(f"  [SKIP] Sheet not found: '{sheet_name}'")
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header_row = rows[0]
    headers = [str(h).strip() if h is not None else None for h in header_row]
    col_map = build_col_map(header_row)

    units = []
    for row in rows[1:]:
        u = map_row(row, col_map, headers)
        if u is None:
            continue
        u["source_file"] = source_file
        u["source_sheet"] = sheet_name
        units.append(u)

    return units


# ── REST API helpers ──────────────────────────────────────────────────────────

def make_session(env: dict) -> tuple:
    """Returns (session, base_url)."""
    base_url = env.get("NEXT_PUBLIC_SUPABASE_URL", "").rstrip("/")
    service_key = env.get("PRIVATE_SUPABASE_SERVICE_KEY", "")
    if not base_url or not service_key:
        raise ValueError("Missing NEXT_PUBLIC_SUPABASE_URL or PRIVATE_SUPABASE_SERVICE_KEY in .env.local")

    session = requests.Session()
    session.headers.update({
        "apikey":        service_key,
        "Authorization": f"Bearer {service_key}",
        "Content-Type":  "application/json",
        "Prefer":        "return=representation",
    })
    return session, f"{base_url}/rest/v1"


def api_check(resp: requests.Response, label: str):
    if not resp.ok:
        raise RuntimeError(f"{label} failed [{resp.status_code}]: {resp.text[:400]}")


def upsert_company(session, base_url: str, name: str, display_name: str) -> str:
    resp = session.post(
        f"{base_url}/real_estate_companies",
        params={"on_conflict": "name"},
        headers={"Prefer": "resolution=merge-duplicates,return=representation"},
        json={"name": name, "display_name": display_name},
    )
    api_check(resp, f"upsert company '{name}'")
    return resp.json()[0]["id"]


def upsert_project(session, base_url: str, company_id: str, proj: dict, source_file: str) -> str:
    resp = session.post(
        f"{base_url}/projects",
        params={"on_conflict": "slug"},
        headers={"Prefer": "resolution=merge-duplicates,return=representation"},
        json={
            "company_id":   company_id,
            "name":         proj["name"],
            "slug":         proj["slug"],
            "commune":      proj["commune"],
            "source_file":  source_file,
            "source_sheet": proj["sheet"],
        },
    )
    api_check(resp, f"upsert project '{proj['slug']}'")
    return resp.json()[0]["id"]


def delete_project_units(session, base_url: str, project_id: str):
    resp = session.delete(
        f"{base_url}/units",
        params={"project_id": f"eq.{project_id}"},
    )
    api_check(resp, f"delete units for project {project_id}")


def insert_units_batch(session, base_url: str, project_id: str, units: list, batch_size: int = 100):
    """Delete existing units then bulk-insert in batches."""
    delete_project_units(session, base_url, project_id)

    for i in range(0, len(units), batch_size):
        batch = units[i: i + batch_size]
        rows = []
        for u in batch:
            row = {
                "project_id":             project_id,
                "unit_number":            u.get("unit_number"),
                "status":                 u.get("status", "blocked"),
                "original_status":        u.get("original_status"),
                "orientation":            u.get("orientation"),
                "typology":               u.get("typology"),
                "unit_type":              u.get("unit_type"),
                "surface_useful":         u.get("surface_useful"),
                "surface_terrace":        u.get("surface_terrace"),
                "surface_weighted":       u.get("surface_weighted"),
                "surface_total":          u.get("surface_total"),
                "list_price":             u.get("list_price"),
                "discount":               u.get("discount"),
                "final_price":            u.get("final_price"),
                "deed_price":             u.get("deed_price"),
                "uf_per_m2":              u.get("uf_per_m2"),
                "parking":                u.get("parking") or 0,
                "storage":                u.get("storage") or 0,
                "bonus_percentage":       u.get("bonus_percentage"),
                "bonus_uf":               u.get("bonus_uf"),
                "mortgage_max_percentage":u.get("mortgage_max_percentage"),
                "mortgage_max_uf":        u.get("mortgage_max_uf"),
                "pie_percentage":         u.get("pie_percentage"),
                "pie_uf":                 u.get("pie_uf"),
                "pie_clp":                u.get("pie_clp"),
                "installments_plan1":     u.get("installments_plan1"),
                "monthly_payment_plan1":  u.get("monthly_payment_plan1"),
                "installments_plan2":     u.get("installments_plan2"),
                "monthly_payment_plan2":  u.get("monthly_payment_plan2"),
                "rent_estimate":          u.get("rent_estimate"),
                "mortgage_payment":       u.get("mortgage_payment"),
                "cash_flow":              u.get("cash_flow"),
                "min_income":             u.get("min_income"),
                "raw_data":               u.get("raw_data", {}),
            }
            rows.append(row)

        resp = session.post(
            f"{base_url}/units",
            headers={"Prefer": "return=minimal"},
            data=json.dumps(rows, ensure_ascii=False, default=str),
        )
        api_check(resp, f"insert units batch {i//batch_size + 1}")


def update_project_counts(session, base_url: str, project_id: str, total: int, available: int):
    resp = session.patch(
        f"{base_url}/projects",
        params={"id": f"eq.{project_id}"},
        headers={"Prefer": "return=minimal"},
        json={"total_units": total, "available_units": available},
    )
    api_check(resp, f"update project counts {project_id}")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("Loading .env.local …")
    env = load_env(ENV_FILE)

    session, base_url = make_session(env)
    print(f"API endpoint: {base_url}")

    # Quick connectivity check
    r = session.get(f"{base_url}/real_estate_companies", params={"limit": "1"})
    if not r.ok:
        print(f"ERROR: Cannot reach API [{r.status_code}]: {r.text[:200]}")
        sys.exit(1)
    print("API connection OK\n")

    total_units = 0

    for company_cfg in COMPANIES:
        company_name = company_cfg["name"]
        file_path = os.path.join(SCRIPT_DIR, company_cfg["file"])

        print(f"{'='*60}")
        print(f"Company: {company_name}  |  {company_cfg['file']}")

        company_id = upsert_company(session, base_url, company_name, company_cfg["display_name"])
        print(f"  company_id = {company_id}")

        wb = openpyxl.load_workbook(file_path, read_only=True, keep_vba=False, data_only=True)

        for proj_cfg in company_cfg["projects"]:
            sheet_name = proj_cfg["sheet"]
            print(f"\n  ── {proj_cfg['name']}  (sheet: {sheet_name})")

            project_id = upsert_project(session, base_url, company_id, proj_cfg, company_cfg["file"])
            print(f"     project_id = {project_id}")

            units = read_sheet(wb, sheet_name, company_cfg["file"])
            print(f"     Read {len(units)} units from Excel")

            if units:
                insert_units_batch(session, base_url, project_id, units)

                status_counts = Counter(u["status"] for u in units)
                available = status_counts.get("available", 0) + status_counts.get("sin_abono", 0)
                update_project_counts(session, base_url, project_id, len(units), available)

                total_units += len(units)
                print(f"     Inserted {len(units)} units  |  {dict(status_counts)}")

        wb.close()
        print()

    print(f"{'='*60}")
    print(f"Done!  Total units inserted: {total_units}")


if __name__ == "__main__":
    main()
